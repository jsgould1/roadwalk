"""
Sign-inventory spreadsheet → sign_reference.json

Converts the field-collected Sign Inventory_10S & 10N.xlsx into a JSON
file the RoadWalk Reference Data panel can load. Each sheet becomes one
"section" of rows; each row is one Post plus one or more Panels. Values
are normalized to match the data form's enum vocab (post_type, foundation,
sheeting_type, etc.) so the in-app "Apply to Form" button can write
directly into pin.attrs without further translation.

Continuation rows (blank "#" column with panel data) are attached as
additional panels to the previous row's post.

Run:
    python data/import_sign_inventory.py "/path/to/Sign Inventory_10S & 10N.xlsx" sign_reference.json

The output JSON file goes wherever you point arg 2 — drop it on the
machine that's running RoadWalk and load it via the Reference Data
panel.
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl


# ── Enum mapping tables ────────────────────────────────────────────────
# Map raw spreadsheet strings to the form's option codes. Anything we
# can't confidently classify is kept in a "_raw" sibling so the user can
# see what the spreadsheet said vs. what auto-fill will write.

POST_TYPE_MAP = {
    "4 x 4 wood":               "WOOD_4X4",
    "4x4 wood":                 "WOOD_4X4",
    "4 x 6 wood":               "WOOD_4X6",
    "4x6 wood":                 "WOOD_4X6",
    "4 x 8 wood":               "WOOD_6X8",   # closest available; flag in notes
    "6 x 6 wood":               "WOOD_6X6",
    "6 x 8 wood":               "WOOD_6X8",
    "u channel steel":          "UCHANNEL",
    "u-channel steel":          "UCHANNEL",
    "u channel":                "UCHANNEL",
    "uchannel":                 "UCHANNEL",
    "round galv pipe":          "PIPE_GALV",
    "galvanized pipe":          "PIPE_GALV",
    "i-beam steel":             "I_BEAM",
    "i beam steel":             "I_BEAM",
    "steel cantelever":         "OTHER",
    "steel cantilever":         "OTHER",
    "2\" square painted steel": "SQ_TUBE_PERF",
    "square tube":              "SQ_TUBE_PERF",
    "aluminum":                 "ALUMINUM",
    "concrete":                 "CONCRETE",
}

FOUNDATION_MAP = {
    "direct bury":      "buried_direct",
    "direct burried":   "buried_direct",
    "directbury":       "buried_direct",
    "direct burry":     "buried_direct",
    "direct buried":    "buried_direct",
    "concrete":         "concrete",
    "concrete footing": "concrete",
    "drive anchor":     "drive_anchor",
    "flange plate":     "flange_plate",
}

BREAKAWAY_MAP = {
    "y": "yes", "yes": "yes",
    "n": "no",  "no":  "no",
    "u": "unknown", "unknown": "unknown",
}

CONDITION_MAP = {
    "e": "E", "excellent": "E",
    "g": "G", "good":      "G",
    "f": "F", "fair":      "F",
    "p": "P", "poor":      "P",
    "x": "X", "fail":      "X", "failed": "X",
}

SHEETING_MAP = {
    "engineer grade":            "EG",
    "engineering grade":         "EG",
    "high intensity prismatic":  "HIP",
    "hexagon":                   "HIP",   # field shorthand for hex-shaped HIP pattern
    "hexagonal":                 "HIP",
    "diamond grade":             "DG3",
    "diamond":                   "DG3",
    "dg3":                       "DG3",
    "fluorescent yellow-green":  "FYG",
    "fluorescent yel-grn":       "FYG",
    "fluorescent orange":        "FO",
    "other":                     "OTHER",
    "unknown":                   "UNKNOWN",
}

BACK_PANEL_MAP = {
    "brown":     "brown",
    "black":     "black",
    "white":     "white",
    "green":     "green",
    "two-sided": "two_sided",
    "two sided": "two_sided",
    "galvanized": "galvanized",
    "galvanized back": "galvanized",
    "galvanized back panel": "galvanized",
    "other":     "other",
}

MOUNTING_MAP = {
    "angle bracket":   "angled_brackets",
    "angled bracket":  "angled_brackets",
    "angled brackets": "angled_brackets",
    "angle brackets":  "angled_brackets",
    "bracket":         "angled_brackets",
    "bolt":            "carriage_bolts",
    "bolted":          "carriage_bolts",
    "bolts":           "carriage_bolts",
    "carriage bolt":   "carriage_bolts",
    "carriage bolts":  "carriage_bolts",
    "screw":           "screws",
    "screws":          "screws",
    "d-clamp":         "d_clamps",
    "d-clamps":        "d_clamps",
    "d clamps":        "d_clamps",
}

# Form's damage_flags enum — substring detector. Anything else stays in
# notes via "damage_text" so nothing's lost.
DAMAGE_FLAGS = {
    "FADED":        ["faded", "fade"],
    "GRAFFITI":     ["graffiti"],
    "STICKERS":     ["sticker"],
    "BULLET_HOLES": ["bullet"],
    "BENT":         ["bent", "dent"],
    "DELAMINATED":  ["delaminat"],
    "CRACKED":      ["crack", "chip", "split"],
    "VEGETATION":   ["vegetation", "overgrown"],
    "MISSING":      ["missing"],
    "DARK_NIGHT":   ["dark at night", "dark night"],
    "SPOTTY_RETRO": ["spotty"],
}

POST_DAMAGE_FLAGS = {
    "RUST":              ["rust", "corrod"],
    "LEAN":              ["lean", "tilt", "loose in ground"],
    "VEHICLE_HIT":       ["impact", "vehicle"],
    "FOUNDATION_EXPOSED": ["foundation exposed", "foundation visible"],
    "MISSING_HARDWARE":  ["missing hardware"],
    "ROT":               ["rot", "rotting", "rotted"],
    "SPLIT":             ["split", "crack"],
    "GRAFFITI":          ["graffiti"],
}


def _norm(v):
    """Trim whitespace, return None for blanks."""
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _lookup(table, raw):
    if raw is None:
        return None, None
    key = str(raw).strip().lower()
    if key in table:
        return table[key], None
    # Try without trailing punctuation
    stripped = re.sub(r"[\s,;:]+$", "", key)
    if stripped in table:
        return table[stripped], None
    return None, str(raw).strip()   # (mapped_value, original_unmapped_text)


def _condition(raw):
    if raw is None:
        return None, None
    s = str(raw).strip()
    if not s:
        return None, None
    # First char often is the code (e.g., "P (One post rotting)")
    first = s[0].upper()
    if first in {"E", "G", "F", "P", "X"}:
        # Did we lose context? Keep the rest as note text.
        rest = s[1:].strip()
        rest = re.sub(r"^[(\s\-:,]+", "", rest).rstrip(") ")
        return first, rest if rest else None
    # Try full-word match
    mapped, raw_text = _lookup(CONDITION_MAP, s)
    return mapped, raw_text


def _detect_flags(raw, flag_table):
    """Return (list_of_codes, leftover_text). Substring match on the lowered raw."""
    if raw is None:
        return [], None
    lo = str(raw).lower()
    matched = []
    for code, needles in flag_table.items():
        if any(n in lo for n in needles):
            matched.append(code)
    return matched, str(raw).strip() if not matched else str(raw).strip()


def _num(v):
    if v is None or v == "":
        return None
    try:
        f = float(v)
        return int(f) if f.is_integer() else f
    except (TypeError, ValueError):
        return None


def _build_post(row):
    """Build the {post:{}, side, eop_offset} block from one row's cells."""
    # Spreadsheet columns (0-based, with col A being blank in every sheet):
    #  B 1  #
    #  C 2  Side of Road
    #  D 3  EOP Offset (ft)
    #  E 4  # of posts
    #  F 5  Post Type
    #  G 6  Height
    #  H 7  Foundation
    #  I 8  Break-away
    #  J 9  Post Condition
    side = _norm(row[2])
    eop  = _num(row[3])
    pt_m, pt_raw   = _lookup(POST_TYPE_MAP, row[5])
    fnd_m, fnd_raw = _lookup(FOUNDATION_MAP, row[7])
    brk_m, brk_raw = _lookup(BREAKAWAY_MAP, row[8])
    cond_v, cond_note = _condition(row[9])
    pdmg, pdmg_raw = _detect_flags(None, POST_DAMAGE_FLAGS)  # no separate post damage col

    post = {
        "post_type":         pt_m,
        "post_type_raw":     pt_raw,
        "post_count":        _num(row[4]),
        "post_height_ft":    _num(row[6]),
        "post_foundation":   fnd_m,
        "post_foundation_raw": fnd_raw,
        "post_breakaway":    brk_m,
        "post_breakaway_raw": brk_raw,
        "post_condition":    cond_v,
        "post_condition_note": cond_note,
        "post_damage":       pdmg,
    }
    return {
        "side":       side,
        "eop_offset": eop,
        "post":       post,
    }


def _build_panel(row, base_col, comment_col):
    """Build one panel record starting at base_col (the Angle column).

    Columns at base_col + offset:
      0  Angle (facing degrees)
      1  Description
      2  Size
      3  Panel Cond
      4  Sheet Cond
      5  Damage
      6  Sheeting Type
      7  Back Panel
      8  Mounting Hardware
      9  Reflectivity Rating
    """
    angle = _num(row[base_col])
    desc  = _norm(row[base_col + 1])
    size  = _norm(row[base_col + 2])
    panel_cond, panel_note   = _condition(row[base_col + 3])
    sheet_cond, sheet_note   = _condition(row[base_col + 4])
    dmg_flags, dmg_raw       = _detect_flags(row[base_col + 5], DAMAGE_FLAGS)
    sht_m, sht_raw           = _lookup(SHEETING_MAP, row[base_col + 6])
    bp_m,  bp_raw            = _lookup(BACK_PANEL_MAP, row[base_col + 7])
    mh_m,  mh_raw            = _lookup(MOUNTING_MAP, row[base_col + 8])
    retro_v, retro_note      = _condition(row[base_col + 9])
    comment = _norm(row[comment_col]) if comment_col is not None and comment_col < len(row) else None

    # If every meaningful field is empty, this panel slot is unused.
    if all(v in (None, [], "") for v in [angle, desc, size, panel_cond, sheet_cond,
                                          dmg_flags, sht_m, bp_m, mh_m, retro_v, comment]):
        return None

    notes_parts = []
    if comment:        notes_parts.append(comment)
    if panel_note:     notes_parts.append("Panel cond note: " + panel_note)
    if sheet_note:     notes_parts.append("Sheet cond note: " + sheet_note)
    if retro_note:     notes_parts.append("Retro note: "      + retro_note)
    if dmg_raw and not dmg_flags:
        notes_parts.append("Damage: " + dmg_raw)
    return {
        "facing_degrees":   angle,
        "panel_description": desc,
        "panel_size":       size,
        "panel_condition":  panel_cond,
        "sheeting_condition": sheet_cond,
        "damage_flags":     dmg_flags,
        "damage_text":      dmg_raw if dmg_raw else None,
        "sheeting_type":    sht_m,
        "sheeting_type_raw": sht_raw,
        "back_panel":       bp_m,
        "back_panel_raw":   bp_raw,
        "mounting_hardware": mh_m,
        "mounting_hardware_raw": mh_raw,
        "retro_day_r1":     retro_v,
        "notes":            " · ".join(notes_parts) if notes_parts else None,
    }


def _read_sheet(ws):
    """Return [{row_idx, row_label, side, eop_offset, post, panels:[…]}, …]."""
    # Locate the header row — first row whose B-column is "#".
    header_row = None
    for ridx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if len(row) > 1 and _norm(row[1]) == "#":
            header_row = ridx
            break
    if header_row is None:
        return []

    # Detect Panel 1 columns: K (10) Angle, L (11) Description, …
    # Panel 2 starts at V (21) Angle for the 10N / 10S MP14-20 sheets.
    # We probe header values to decide.
    hdr = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
    panel_blocks = []
    # Block #1 is always K..U (cols 10..20).
    if len(hdr) > 20 and _norm(hdr[10]) == "Angle":
        panel_blocks.append((10, 20))   # (base, comment_col)
    # Block #2 if a second "Angle" header is present further right.
    if len(hdr) > 30 and _norm(hdr[21]) == "Angle":
        panel_blocks.append((21, 31 if len(hdr) > 31 and _norm(hdr[31]) == "Comment" else None))
    # Block #3 only if columns extend that far.
    if len(hdr) > 41 and _norm(hdr[32]) == "Angle":
        panel_blocks.append((32, 42 if len(hdr) > 42 and _norm(hdr[42]) == "Comment" else None))

    out = []
    last_record = None
    for ridx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
        if len(row) < 2:
            continue
        # Pad row to at least the columns the panel blocks need.
        max_needed = max((b[0] + 10 for b in panel_blocks), default=21)
        if len(row) < max_needed + 1:
            row = list(row) + [None] * (max_needed + 1 - len(row))

        post_num = _num(row[1])
        # Continuation row — attach panel to the last record we built.
        if post_num is None:
            if last_record is None:
                continue
            for base, comment_col in panel_blocks:
                panel = _build_panel(row, base, comment_col)
                if panel:
                    last_record["panels"].append(panel)
            continue

        # Fresh record.
        record = _build_post(row)
        record["row_idx"]   = post_num
        record["row_label"] = f"#{post_num}"
        record["panels"]    = []
        for base, comment_col in panel_blocks:
            panel = _build_panel(row, base, comment_col)
            if panel:
                record["panels"].append(panel)
        out.append(record)
        last_record = record
    return out


def main(in_path: Path, out_path: Path) -> None:
    wb = openpyxl.load_workbook(in_path, data_only=True)
    sheets_out = {}
    total_rows = 0
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = _read_sheet(ws)
        sheets_out[sn] = {
            "label":     sn,
            "row_count": len(rows),
            "rows":      rows,
        }
        total_rows += len(rows)

    payload = {
        "_format":      "roadwalk-sign-reference-v1",
        "source_file":  in_path.name,
        "sheets":       sheets_out,
        "total_rows":   total_rows,
    }
    out_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Wrote {out_path} — {total_rows} rows across {len(sheets_out)} sheet(s)")
    for sn, meta in sheets_out.items():
        print(f"  {sn:25s}  {meta['row_count']} row(s)")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: import_sign_inventory.py INPUT.xlsx OUTPUT.json")
        sys.exit(2)
    main(Path(sys.argv[1]), Path(sys.argv[2]))
