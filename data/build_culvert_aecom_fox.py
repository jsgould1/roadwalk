"""Build the AECOM_FOX culvert bundle for RoadWalk's Section A.

Inputs (read-only — none are modified):
  - GRSM_Culvert 1.kmz     - 60 LineString culverts with full attributes
  - Copy of GRSM_Culverts.xlsx - same rows + Picture-in-Cell inlet/outlet photos
  - section-A-gatlinburg-bypass.geojson - alignment used to compute sta_ft / side

Output:
  - data/culvert-aecom-fox-bundle.json - culvert pins ready to merge into
    prewalk-bundle.json under sections[A].pins

Photos are embedded as data: URIs on each pin's attrs.photos so the bundle
loads in one shot with no separate photo folder to manage. Resized to
1200 px wide JPEG quality 0.78 - a 60-culvert × 2-photo run ends up at ~10 MB.

The KMZ "Num" field equals the Excel "Order" column - confirmed by matching
Notes between rows (Order 5 / Num 5 both carry "Impaired"). Excel rows
without a Num match drop out of the bundle and get logged.

Run from the RoadWalk root:
    python data/build_culvert_aecom_fox.py
"""

from __future__ import annotations

import base64
import io
import json
import math
import re
import sys
import zipfile
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from PIL import Image  # Pillow ships with most CPython installs at AECOM

ROOT       = Path(__file__).resolve().parent
KMZ_PATH   = Path(r"C:\Users\gouldj\Downloads\GRSM_Culvert 1.kmz")
XLSX_PATH  = Path(r"C:\Users\gouldj\OneDrive - AECOM\Documents\!DATA\EFL\2026 GRSM\Copy of GRSM_Culverts.xlsx")
ALIGN_GEOJSON = ROOT / "section-A-gatlinburg-bypass.geojson"
OUT_BUNDLE = ROOT / "culvert-aecom-fox-bundle.json"

SECTION_ID = "A"
SOURCE_TAG = "AECOM_FOX"          # marks every pin so the report can filter
PHOTO_MAX_WIDTH = 1200
PHOTO_JPEG_QUALITY = 78


# ── KMZ parser ─────────────────────────────────────────────────────────────
def parse_kmz(path: Path) -> list[dict]:
    """Pull every Placemark from doc.kml. Each entry returns:
       { num: int, coords: [[lng,lat], [lng,lat], ...], attrs: {kml column → value} }
    """
    with zipfile.ZipFile(path) as z:
        kml = z.read("doc.kml").decode("utf-8", errors="ignore")

    out: list[dict] = []
    for pm in re.findall(r"<Placemark[^>]*>(.*?)</Placemark>", kml, re.DOTALL):
        # Coordinate list inside the (Multi)Geometry block
        coord_m = re.search(r"<coordinates>\s*(.*?)\s*</coordinates>", pm, re.DOTALL)
        if not coord_m:
            continue
        coords: list[list[float]] = []
        for raw in coord_m.group(1).split():
            parts = raw.split(",")
            if len(parts) < 2:
                continue
            try:
                lng, lat = float(parts[0]), float(parts[1])
            except ValueError:
                continue
            coords.append([lng, lat])
        if len(coords) < 2:
            continue

        # Extract the attribute table from the CDATA description
        attrs: dict[str, str] = {}
        for label, value in re.findall(
            r"<td>([^<]+?)</td>\s*<td>([^<]*)</td>", pm
        ):
            attrs[label.strip()] = value.strip()

        try:
            num = int(attrs.get("Num", "") or attrs.get("FID", ""))
        except ValueError:
            continue
        out.append({"num": num, "coords": coords, "attrs": attrs})
    return out


# ── Excel parser (with Picture-in-Cell image extraction) ───────────────────
def _parse_namespaced_attr(elem: str, attr: str) -> str | None:
    m = re.search(rf'{attr}="([^"]*)"', elem)
    return m.group(1) if m else None


def parse_xlsx_with_pics(path: Path) -> dict[int, dict]:
    """Return { Order: { excel cols dict, inlet_bytes?, outlet_bytes? } }.

    Excel's "Picture in Cell" mechanism (Excel 365) leaves the cell value
    as #VALUE! to anything that isn't Excel itself - the image lives on
    a chain of rich-value metadata:

        cell vm="K"
          -> valueMetadata bk[K-1]  (futureMetadata index N)
            -> futureMetadata XLRICHVALUE bk[N]  (rich value index R)
              -> richValueRel.xml rels[R]  (rId)
                -> richValueRel rels file (rId → media/image.png)

    We parse it manually rather than relying on openpyxl, which doesn't
    support Picture in Cell yet.
    """
    wb = load_workbook(path, data_only=False)
    ws = wb["Table"]
    header = [c.value for c in ws[1]]

    # Locate the Order, Inlet, Outlet columns by header text
    col_idx = {h: i for i, h in enumerate(header)}
    order_col  = col_idx["Order"] + 1
    inlet_col  = col_idx["Inlet "] + 1   # trailing space matches the file
    outlet_col = col_idx["Outlet"] + 1

    # Pull every cell value (including formulas/None) + vm attribute for the
    # inlet/outlet columns. openpyxl exposes value but not vm, so re-parse
    # the raw sheet XML to grab them.
    with zipfile.ZipFile(path) as z:
        sheet_xml = z.read("xl/worksheets/sheet1.xml").decode("utf-8")
        meta_xml  = z.read("xl/metadata.xml").decode("utf-8")
        rv_xml    = z.read("xl/richData/richValueRel.xml").decode("utf-8")
        rv_rels_xml = z.read("xl/richData/_rels/richValueRel.xml.rels").decode("utf-8")

        # Build the rId → image file path map. Use a lazy `.+?` instead of
        # `[^/]*` because the Target attribute carries paths like
        # "../media/image117.png" — the literal slash would otherwise abort
        # the match and leave the dict empty.
        rid_to_image: dict[str, str] = {}
        for rel in re.findall(r"<Relationship\s+[^>]+?/>", rv_rels_xml):
            rid = _parse_namespaced_attr(rel, "Id")
            target = _parse_namespaced_attr(rel, "Target")
            if rid and target:
                rid_to_image[rid] = target.lstrip("./")
        # richValueRel.xml: ordered list of <rel r:id="rIdN"/>; index ⇒ rId
        rv_index_to_rid = [
            _parse_namespaced_attr(rel, "r:id")
            for rel in re.findall(r"<rel [^/]*/>", rv_xml)
        ]

        # futureMetadata XLRICHVALUE: index ⇒ rich value index (xlrd:rvb i)
        # Pull each <bk>…</bk> in the order they appear inside the
        # name="XLRICHVALUE" block.
        fm_block_m = re.search(
            r'<futureMetadata name="XLRICHVALUE"[^>]*>(.*?)</futureMetadata>',
            meta_xml, re.DOTALL,
        )
        fm_rv_index: list[int] = []
        if fm_block_m:
            for bk in re.findall(r"<bk>(.*?)</bk>", fm_block_m.group(1), re.DOTALL):
                i_m = re.search(r'i="(\d+)"', bk)
                fm_rv_index.append(int(i_m.group(1)) if i_m else -1)

        # valueMetadata: index ⇒ futureMetadata index (rc t="1" v="N")
        vm_block_m = re.search(
            r"<valueMetadata[^>]*>(.*?)</valueMetadata>", meta_xml, re.DOTALL
        )
        vm_fm_index: list[int] = []
        if vm_block_m:
            for bk in re.findall(r"<bk>(.*?)</bk>", vm_block_m.group(1), re.DOTALL):
                m = re.search(r'v="(\d+)"', bk)
                vm_fm_index.append(int(m.group(1)) if m else -1)

        def vm_to_image_path(vm: int) -> str | None:
            """Resolve a cell vm="K" attribute to an xl/media/imageN.png path."""
            try:
                fm = vm_fm_index[vm - 1]                 # vm is 1-based
                rv = fm_rv_index[fm]
                rid = rv_index_to_rid[rv]
                img = rid_to_image[rid]
            except (IndexError, KeyError):
                return None
            return "xl/" + img

        # Pre-extract every cell's vm attribute for the inlet/outlet columns.
        # Sheet XML cell tag looks like:
        #   <c r="M2" s="2" t="e" vm="1"><v>#VALUE!</v></c>
        # Match the opening tag only - matching the closing </c> with an
        # arbitrary inner-content pattern misses cells whose inner shape
        # doesn't fit the pattern, which is how we got zero hits last run.
        cell_vm_map: dict[str, int] = {
            r: int(vm)
            for r, vm in re.findall(
                r'<c r="([A-Z]+\d+)"[^>]*vm="(\d+)"', sheet_xml
            )
        }

        # Now walk rows and build the output dict
        col_letter = {1: "A", 2: "B", 3: "C", 4: "D", 5: "E", 6: "F",
                      7: "G", 8: "H", 9: "I", 10: "J", 11: "K", 12: "L",
                      13: "M", 14: "N"}

        def load_image_bytes(image_path: str | None) -> bytes | None:
            if not image_path:
                return None
            try:
                return z.read(image_path)
            except KeyError:
                return None

        rows: dict[int, dict] = {}
        for r in range(2, ws.max_row + 1):
            row_vals = {h: ws.cell(r, c + 1).value for c, h in enumerate(header)}
            order = row_vals.get("Order")
            if not isinstance(order, int):
                # Some Excel rows are blank/header-repeats - skip
                continue

            inlet_vm  = cell_vm_map.get(f"{col_letter[inlet_col]}{r}")
            outlet_vm = cell_vm_map.get(f"{col_letter[outlet_col]}{r}")
            inlet_bytes  = load_image_bytes(vm_to_image_path(inlet_vm))  if inlet_vm  else None
            outlet_bytes = load_image_bytes(vm_to_image_path(outlet_vm)) if outlet_vm else None

            rows[order] = {
                "row": row_vals,
                "inlet_bytes":  inlet_bytes,
                "outlet_bytes": outlet_bytes,
            }
    return rows


# ── Section A alignment + stationing projection ────────────────────────────
def load_alignment(path: Path) -> list[tuple[float, float]]:
    """Returns the alignment as a list of (lat, lng) pairs."""
    geo = json.loads(path.read_text(encoding="utf-8"))
    feats = geo.get("features", [])
    if not feats:
        raise SystemExit(f"No features in {path}")
    coords = feats[0]["geometry"]["coordinates"]
    # GeoJSON LineString → [[lng,lat], ...]; return (lat, lng) for downstream
    return [(c[1], c[0]) for c in coords]


def haversine_ft(a: tuple[float, float], b: tuple[float, float]) -> float:
    """Distance in feet between two (lat, lng) points."""
    R_FT = 20902231.0  # Earth radius in feet
    lat1, lng1 = a
    lat2, lng2 = b
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlam = math.radians(lng2 - lng1)
    h = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlam / 2) ** 2
    return 2 * R_FT * math.asin(math.sqrt(h))


def project_pt_onto_alignment(
    pt: tuple[float, float], align: list[tuple[float, float]]
) -> tuple[float, float, str]:
    """Returns (sta_ft, perp_offset_ft, side) where side is 'L', 'R', or 'C'.

    Equirectangular approximation per segment; fine at section scale.
    """
    best: tuple[float, float, str] | None = None
    best_d2 = float("inf")
    cum_ft = 0.0

    lat0 = sum(p[0] for p in align) / len(align)
    coslat0 = math.cos(math.radians(lat0))
    FT_PER_DEG = 364000.0  # ~1 deg ≈ 364 000 ft on Earth's surface

    px, py = pt[1] * coslat0 * FT_PER_DEG, pt[0] * FT_PER_DEG

    for i in range(len(align) - 1):
        a, b = align[i], align[i + 1]
        ax, ay = a[1] * coslat0 * FT_PER_DEG, a[0] * FT_PER_DEG
        bx, by = b[1] * coslat0 * FT_PER_DEG, b[0] * FT_PER_DEG
        seg_dx, seg_dy = bx - ax, by - ay
        seg_len2 = seg_dx * seg_dx + seg_dy * seg_dy
        if seg_len2 == 0:
            continue
        t = ((px - ax) * seg_dx + (py - ay) * seg_dy) / seg_len2
        t_clamped = max(0.0, min(1.0, t))
        proj_x = ax + t_clamped * seg_dx
        proj_y = ay + t_clamped * seg_dy
        d2 = (px - proj_x) ** 2 + (py - proj_y) ** 2
        if d2 < best_d2:
            seg_len = math.sqrt(seg_len2)
            sta_ft = cum_ft + t_clamped * seg_len
            # Signed cross product → side (positive = left of forward bearing)
            cross = seg_dx * (py - ay) - seg_dy * (px - ax)
            side = "L" if cross > 0 else "R" if cross < 0 else "C"
            offset = math.sqrt(d2)
            best = (sta_ft, offset, side)
            best_d2 = d2
        cum_ft += math.sqrt(seg_len2)

    return best if best else (0.0, 0.0, "C")


def fmt_sta(sta_ft: float) -> str:
    """Convert 877.0 → '8+77'."""
    sta_ft = max(0.0, sta_ft)
    full = int(sta_ft / 100)
    rem  = sta_ft - full * 100
    return f"{full}+{rem:02.0f}"


# ── Photo compression ──────────────────────────────────────────────────────
def png_bytes_to_jpeg_dataurl(png_bytes: bytes, max_width: int, quality: int) -> str:
    """Compress + resize source PNG bytes to a JPEG data URL."""
    img = Image.open(io.BytesIO(png_bytes))
    if img.mode != "RGB":
        img = img.convert("RGB")
    if img.width > max_width:
        ratio = max_width / img.width
        img = img.resize((max_width, int(img.height * ratio)), Image.LANCZOS)
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=quality, optimize=True)
    b64 = base64.b64encode(buf.getvalue()).decode("ascii")
    return f"data:image/jpeg;base64,{b64}"


# ── Main bundle build ──────────────────────────────────────────────────────
def safe_int(v) -> int | None:
    try:
        return int(v) if v not in (None, "") else None
    except (ValueError, TypeError):
        return None


def safe_str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def build_bundle() -> dict:
    print(f"Reading KMZ from {KMZ_PATH}")
    kmz_culverts = parse_kmz(KMZ_PATH)
    print(f"  {len(kmz_culverts)} placemarks")

    print(f"Reading Excel from {XLSX_PATH}")
    excel_by_order = parse_xlsx_with_pics(XLSX_PATH)
    print(f"  {len(excel_by_order)} rows by Order")
    with_inlet  = sum(1 for v in excel_by_order.values() if v["inlet_bytes"])
    with_outlet = sum(1 for v in excel_by_order.values() if v["outlet_bytes"])
    print(f"  inlet photos resolved: {with_inlet} / outlet: {with_outlet}")

    print(f"Loading Section A alignment from {ALIGN_GEOJSON}")
    align = load_alignment(ALIGN_GEOJSON)
    print(f"  {len(align)} vertices")

    pins: list[dict] = []
    skipped: list[str] = []

    for kc in kmz_culverts:
        order = kc["num"]
        excel_row = excel_by_order.get(order)
        if not excel_row:
            skipped.append(f"KMZ Num {order} - no matching Excel Order")
            continue

        row = excel_row["row"]
        coords = kc["coords"]
        # Centroid + endpoints used for stationing + side calc
        c_lng = sum(p[0] for p in coords) / len(coords)
        c_lat = sum(p[1] for p in coords) / len(coords)
        sta_ft, offset_ft, side = project_pt_onto_alignment(
            (c_lat, c_lng), align
        )

        # Build photo list — embed both inlet + outlet at full quality with
        # captions so the report card photo grid renders them in order.
        photos: list[dict] = []
        if excel_row["inlet_bytes"]:
            try:
                photos.append({
                    "dataUrl": png_bytes_to_jpeg_dataurl(
                        excel_row["inlet_bytes"], PHOTO_MAX_WIDTH, PHOTO_JPEG_QUALITY
                    ),
                    "caption": f"Inlet — Culvert {order}",
                    "_role": "inlet",
                })
            except Exception as exc:
                skipped.append(f"Culvert {order} inlet photo: {exc}")
        if excel_row["outlet_bytes"]:
            try:
                photos.append({
                    "dataUrl": png_bytes_to_jpeg_dataurl(
                        excel_row["outlet_bytes"], PHOTO_MAX_WIDTH, PHOTO_JPEG_QUALITY
                    ),
                    "caption": f"Outlet — Culvert {order}",
                    "_role": "outlet",
                })
            except Exception as exc:
                skipped.append(f"Culvert {order} outlet photo: {exc}")

        pin_id = f"CV-{order:03d}"
        # Endpoints stamped so the existing Leaflet drag/relocate code keeps
        # both ends of the line in sync.
        attrs = {
            "source":          SOURCE_TAG,
            "name":            f"Culvert {order}",
            "rpt_type":        "culvert",
            "endpoint_1":      coords[0],
            "endpoint_2":      coords[-1],
            "fox_order":       order,
            "fox_id":          safe_int(row.get("Id")),
            "fox_oid":         safe_int(row.get("OID_")),
            "in_type":         safe_str(row.get("In_Type")),
            "out_type":        safe_str(row.get("Out_Type")),
            "material":        safe_str(row.get("Material")),
            "coating":         safe_str(row.get("Coating")),
            "length_ft":       safe_int(row.get("Length (ft)")),
            "size_in":         safe_int(row.get("Size (in)")),
            "notes":           safe_str(row.get("Notes")),
            "stationing_2026": safe_str(row.get("Stationing")),
            "stationing_1994": safe_str(row.get("Pla94Sta")),
            "perp_offset_ft":  round(offset_ft, 1),
            "perp_side":       side,
            "photos":          photos,
        }
        # Computed station label so the SLD shows a value even if the field
        # data form is never opened.
        attrs["sta_formatted"] = fmt_sta(sta_ft)

        pins.append({
            "id":       pin_id,
            "kind":     "culvert",
            "source":   SOURCE_TAG,
            "status":   "active",
            "geometry": {"type": "LineString", "coordinates": coords},
            "sta_ft":   round(sta_ft, 2),
            "sta":      fmt_sta(sta_ft),
            "attrs":    attrs,
        })

    pins.sort(key=lambda p: p["sta_ft"])

    bundle = {
        "version":      1,
        "source":       SOURCE_TAG,
        "section_id":   SECTION_ID,
        "generated_at": "2026-06-24",
        "pin_count":    len(pins),
        "skipped":      skipped,
        "pins":         pins,
    }
    return bundle


def main() -> int:
    bundle = build_bundle()
    OUT_BUNDLE.write_text(json.dumps(bundle, separators=(",", ":")), encoding="utf-8")
    size_mb = OUT_BUNDLE.stat().st_size / (1024 * 1024)
    print()
    print(f"Wrote {OUT_BUNDLE.relative_to(ROOT.parent)} - {bundle['pin_count']} pins, {size_mb:.1f} MB")
    if bundle["skipped"]:
        print("Skipped:")
        for s in bundle["skipped"]:
            print(" ", s)
    return 0


if __name__ == "__main__":
    sys.exit(main())
