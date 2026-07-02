"""Rebuild the bundle's culvert pins from:
  1. Steven's corrected SF QC KML (geometry + Num)   ← locations + culvert No.
  2. The JSG-updated XLSX (attrs keyed by Order)     ← everything else

Join key: KML <Num> matches XLSX Order as a string, so "21A" and "44A"
line up along with the 1..59 numeric Nums.

Preserves per-pin from the existing bundle:
  - ulid                              (immutable identifier)
  - attrs.photos                      (manifest-imported photos)
  - attrs.report_hidden_photos        (per-photo hide selections)
  - attrs.<x>_condition + _notes      (inlet / barrel / outlet /
                                       hw_upstream / hw_downstream)
  - attrs._user_edited                (dirty flag from inline edits)
  - Geometry-block attrs (bearing / elevations / drop / slope / skew)
                                      — computed downstream by
                                      build_flow_geometry.py; this
                                      script doesn't touch them.

Overwrites per-pin from the new source:
  - id                                (XLSX Culvert ID, e.g. A-CV-005)
  - geometry                          (KML LineString)
  - sta_ft, sta                       (re-projected onto section alignment)
  - attrs.material, coating, in_type, in_side, out_type, out_side
  - attrs.length_ft, size_in
  - attrs.drainage_basin_ac, drainage_direction
  - attrs.notes, stationing_label, pla94_sta_label
  - attrs.survey_num                  ← the KML Num (e.g. "5", "21A")

Run from anywhere:
    python data/import_culvert_kml_and_xlsx.py

Paths at the top of the file are hard-coded to the user's current
KML + XLSX; adjust if a future revision lives elsewhere.
"""

from __future__ import annotations

import json
import math
import re
import sys
import zipfile
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
BUNDLE_PATH = ROOT / "data" / "prewalk-bundle.json"
KML_PATH = (
    Path.home()
    / "Downloads" / "GRSM_Culvert_SF_QC_corrected_2026-07-01-21-06-09.kml"
)
XLSX_PATH = (
    Path.home()
    / "OneDrive - AECOM" / "Documents" / "!DATA" / "EFL" / "2026 GRSM" / "GRSM_Culvert"
    / "GRSM_Culvert_Stationing_SF+JSG 070126 (original phtoos removed).xlsx"
)

SECTIONS_OF_INTEREST = {"A", "G", "H", "I"}

# ── Attrs the new source is authoritative for. Everything else on
# ── the pin's attrs dict is preserved verbatim across the rebuild.
XLSX_OWNED_ATTRS = {
    "material", "coating",
    "in_type", "in_side", "out_type", "out_side",
    "length_ft", "size_in",
    "drainage_basin_ac", "drainage_direction",
    "notes", "stationing_label", "pla94_sta_label",
    "survey_num",
}

# ── ULID helpers (crockford base32 — same format the app uses) ────
_ULID_ALPHABET = "0123456789ABCDEFGHJKMNPQRSTVWXYZ"


def _new_ulid() -> str:
    import os, time
    ts = int(time.time() * 1000)
    ts_chars = []
    for _ in range(10):
        ts_chars.append(_ULID_ALPHABET[ts & 0x1F]); ts >>= 5
    rand = os.urandom(10)
    rand_int = int.from_bytes(rand, "big")
    r_chars = []
    for _ in range(16):
        r_chars.append(_ULID_ALPHABET[rand_int & 0x1F]); rand_int >>= 5
    return "".join(reversed(ts_chars)) + "".join(reversed(r_chars))


# ── Geometry helpers ─────────────────────────────────────────────
M2FT = 3.280839895
R_EARTH_M = 6371000.0


def _hav_ft(a_lat: float, a_lon: float, b_lat: float, b_lon: float) -> float:
    p1 = math.radians(a_lat); p2 = math.radians(b_lat)
    dp = math.radians(b_lat - a_lat); dl = math.radians(b_lon - a_lon)
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * R_EARTH_M * math.asin(math.sqrt(a)) * M2FT


def _project_onto_alignment(pt_lat: float, pt_lon: float,
                            alignment: list) -> tuple[float, float]:
    """Return (sta_ft along alignment at the perpendicular foot,
    perpendicular distance in ft). alignment is [[lng, lat], ...]
    (GeoJSON order — matches how RoadWalk's bundle stores sec.alignment)."""
    if not alignment or len(alignment) < 2:
        return (0.0, float("inf"))
    best_sta = 0.0; best_perp = float("inf"); cum = 0.0
    for i in range(len(alignment) - 1):
        a_lon, a_lat = alignment[i][0],     alignment[i][1]
        b_lon, b_lat = alignment[i + 1][0], alignment[i + 1][1]
        seg_len_ft = _hav_ft(a_lat, a_lon, b_lat, b_lon)
        if seg_len_ft > 0:
            # Small-angle flat approx around the segment midpoint.
            mid_lat = (a_lat + b_lat) / 2
            deg_per_ft_lat = 1 / (69.0 * 5280)
            deg_per_ft_lon = 1 / (69.0 * 5280 * math.cos(math.radians(mid_lat)))
            ax = (a_lon - b_lon) / deg_per_ft_lon
            ay = (a_lat - b_lat) / deg_per_ft_lat
            px = (pt_lon - b_lon) / deg_per_ft_lon
            py = (pt_lat - b_lat) / deg_per_ft_lat
            L2 = ax * ax + ay * ay
            t = 0.0 if L2 == 0 else max(0.0, min(1.0, (px * ax + py * ay) / L2))
            proj_x = b_lon + (a_lon - b_lon) * t
            proj_y = b_lat + (a_lat - b_lat) * t
            perp_ft = _hav_ft(pt_lat, pt_lon, proj_y, proj_x)
            if perp_ft < best_perp:
                best_perp = perp_ft
                best_sta = cum + seg_len_ft * (1 - t)
        cum += seg_len_ft
    return (best_sta, best_perp)


def _fmt_sta(sta_ft: float) -> str:
    sta_ft = max(0.0, sta_ft)
    hundreds = int(sta_ft // 100)
    remainder = sta_ft - hundreds * 100
    return f"{hundreds}+{remainder:05.2f}"


# ── KML parsing ──────────────────────────────────────────────────
_PLACEMARK_RE = re.compile(r"<Placemark>[\s\S]*?</Placemark>")
_NAME_RE = re.compile(r"<name>([^<]+)</name>")
_DATA_RE = re.compile(r'<Data name="([^"]+)"><value>([^<]*)</value></Data>')
_COORDS_RE = re.compile(r"<coordinates>([\s\S]*?)</coordinates>")


def _read_kml_or_kmz(path: Path) -> str:
    if path.suffix.lower() == ".kmz":
        with zipfile.ZipFile(path) as z:
            kml_name = next((n for n in z.namelist() if n.lower().endswith(".kml")), None)
            if not kml_name:
                raise RuntimeError(f"{path.name}: no .kml inside KMZ")
            return z.read(kml_name).decode("utf-8", errors="replace")
    return path.read_text(encoding="utf-8")


def _parse_kml(path: Path) -> list[dict]:
    text = _read_kml_or_kmz(path)
    features = []
    for m in _PLACEMARK_RE.finditer(text):
        body = m.group(0)
        name_m = _NAME_RE.search(body); coords_m = _COORDS_RE.search(body)
        if not coords_m:
            continue
        # Coords: "lon,lat,alt lon,lat,alt ..."
        coords = []
        for tok in coords_m.group(1).split():
            parts = tok.split(",")
            if len(parts) < 2:
                continue
            try:
                lon = float(parts[0]); lat = float(parts[1])
            except ValueError:
                continue
            coords.append([lon, lat])
        if len(coords) < 2:
            continue
        attrs = {k: v for k, v in _DATA_RE.findall(body)}
        name = (name_m.group(1).strip() if name_m else "").strip()
        num = attrs.get("Num", name).strip()
        features.append({
            "num": num,
            "name": name,
            "coords": coords,
            "attrs": attrs,
        })
    return features


# ── XLSX parsing ─────────────────────────────────────────────────
COL = {
    "section_id": 0, "culvert_id": 1, "order": 2,
    "in_type": 3, "in_side": 4, "out_type": 5, "out_side": 6,
    "material": 7, "length_ft": 8, "size_in": 9,
    "drainage_basin_ac": 10, "drainage_dir": 11,
    "notes": 12, "stationing": 13, "pla94_sta": 14,
}


def _str(v) -> str:
    if v is None:
        return ""
    # Openpyxl sometimes returns integer-valued floats (58.0 for a
    # cell formatted as General with value 58). Strip the ".0" so the
    # string round-trips against the KML's plain "58".
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    s = str(v).strip()
    if s in ("None", "nan", ""):
        return ""
    return s


def _fnum(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _read_xlsx(path: Path) -> dict[str, dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    by_order: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[COL["order"]] is None:
            continue
        order = _str(row[COL["order"]])
        if not order:
            continue
        # Section ID cell is occasionally blank (e.g. Order 58 in the
        # 070126 XLSX); fall back to the Culvert ID's leading prefix
        # so we still know where the pin belongs.
        sec_id = _str(row[COL["section_id"]])
        cid    = _str(row[COL["culvert_id"]])
        if not sec_id and "-" in cid:
            sec_id = cid.split("-", 1)[0]
        by_order[order] = {
            "section_id": sec_id,
            "culvert_id": cid,
            "in_type":  _str(row[COL["in_type"]]),
            "in_side":  _str(row[COL["in_side"]]),
            "out_type": _str(row[COL["out_type"]]),
            "out_side": _str(row[COL["out_side"]]),
            "material": _str(row[COL["material"]]),
            "length_ft": _fnum(row[COL["length_ft"]]),
            "size_in":   _fnum(row[COL["size_in"]]),
            "drainage_basin_ac": _fnum(row[COL["drainage_basin_ac"]]),
            "drainage_direction": _str(row[COL["drainage_dir"]]),
            "notes": _str(row[COL["notes"]]),
            "stationing_label": _str(row[COL["stationing"]]),
            "pla94_sta_label": _str(row[COL["pla94_sta"]]),
        }
    return by_order


# ── Main ─────────────────────────────────────────────────────────
def main() -> int:
    if not KML_PATH.exists():
        print(f"ERROR: KML not found at {KML_PATH}", file=sys.stderr); return 1
    if not XLSX_PATH.exists():
        print(f"ERROR: XLSX not found at {XLSX_PATH}", file=sys.stderr); return 1
    if not BUNDLE_PATH.exists():
        print(f"ERROR: bundle not found at {BUNDLE_PATH}", file=sys.stderr); return 1

    print(f"Reading KML   : {KML_PATH}")
    kml_features = _parse_kml(KML_PATH)
    print(f"  {len(kml_features)} Placemark(s)")

    print(f"Reading XLSX  : {XLSX_PATH}")
    xlsx = _read_xlsx(XLSX_PATH)
    print(f"  {len(xlsx)} Order-keyed row(s)")

    print(f"Reading bundle: {BUNDLE_PATH}")
    bundle = json.loads(BUNDLE_PATH.read_text(encoding="utf-8"))
    sections_by_id = {s["id"]: s for s in bundle.get("sections", [])
                      if s.get("id") in SECTIONS_OF_INTEREST}

    # Deduplicate bundle pins by (sec_id, id) before we index them.
    # The bundle sometimes carries multiple copies of the same culvert
    # id (leftover from earlier applyBundle Pass 2 runs). Keep the one
    # with the most inspection data so we don't clobber the field
    # crew's photos + conditions during the rebuild.
    def _pin_score(p: dict) -> int:
        a = p.get("attrs") or {}
        s = 0
        s += len(a.get("photos") or []) * 100
        s += len(a.get("report_hidden_photos") or []) * 50
        for k in ("inlet_condition", "barrel_condition", "outlet_condition",
                  "hw_upstream_cond", "hw_downstream_cond"):
            if a.get(k): s += 10
        for k in ("inlet_condition_notes", "barrel_condition_notes",
                  "outlet_condition_notes", "hw_upstream_cond_notes",
                  "hw_downstream_cond_notes"):
            if a.get(k): s += 5
        if a.get("notes"): s += 3
        return s
    deduped_count = 0
    for sec_id, sec in sections_by_id.items():
        pins = sec.get("pins") or []
        best_by_key: dict[tuple[str, str], dict] = {}
        others: list[dict] = []
        for pin in pins:
            if (pin.get("kind") or "") != "culvert":
                others.append(pin); continue
            k = (sec_id, pin.get("id") or "")
            prior = best_by_key.get(k)
            if prior is None:
                best_by_key[k] = pin
            else:
                if _pin_score(pin) > _pin_score(prior):
                    best_by_key[k] = pin
                deduped_count += 1
        sec["pins"] = others + list(best_by_key.values())

    # Index existing bundle pins by (sec_id, culvert_id) so we can
    # preserve ULID + inspection data across the rebuild.
    existing_by_sec_id: dict[tuple[str, str], dict] = {}
    for sec_id, sec in sections_by_id.items():
        for pin in sec.get("pins") or []:
            if (pin.get("kind") or "") != "culvert":
                continue
            key = (sec_id, pin.get("id") or "")
            existing_by_sec_id[key] = pin

    # ── Join KML → XLSX by Num string ───────────────────────────
    dup_nums: dict[str, int] = {}
    for f in kml_features:
        dup_nums[f["num"]] = dup_nums.get(f["num"], 0) + 1
    dupes = {k: v for k, v in dup_nums.items() if v > 1}
    if dupes:
        print(f"\n⚠ KML has duplicate Num values (kept first, dropped rest): {dupes}")

    consumed = set()
    matched: list[dict] = []                            # {kml, xlsx, sec_id, culvert_id}
    kml_only: list[str] = []
    for f in kml_features:
        num = f["num"]
        if num in consumed:
            continue                                    # duplicate — skip
        consumed.add(num)
        xrow = xlsx.get(num)
        if not xrow:
            kml_only.append(num); continue
        sec_id = xrow["section_id"] or ""
        if sec_id not in SECTIONS_OF_INTEREST:
            kml_only.append(num); continue
        matched.append({
            "kml": f, "xlsx": xrow,
            "sec_id": sec_id,
            "culvert_id": xrow["culvert_id"] or f"{sec_id}-CV-{num}",
        })

    xlsx_only = sorted(set(xlsx.keys()) - consumed)

    # ── Apply matches ───────────────────────────────────────────
    updated = 0; created = 0
    seen_pin_keys: set[tuple[str, str]] = set()
    for m in matched:
        sec_id = m["sec_id"]; culvert_id = m["culvert_id"]
        sec = sections_by_id.get(sec_id)
        if not sec:
            continue
        seen_pin_keys.add((sec_id, culvert_id))
        # Midpoint for sta projection.
        coords = m["kml"]["coords"]
        mid_lat = sum(c[1] for c in coords) / len(coords)
        mid_lon = sum(c[0] for c in coords) / len(coords)
        alignment = sec.get("alignment") or []
        sta_ft, _perp = _project_onto_alignment(mid_lat, mid_lon, alignment)
        # Preserve or create pin.
        pin = existing_by_sec_id.get((sec_id, culvert_id))
        if pin is None:
            pin = {
                "ulid": _new_ulid(),
                "kind": "culvert",
                "id":   culvert_id,
                "attrs": {},
            }
            sec.setdefault("pins", []).append(pin); created += 1
        else:
            updated += 1
        # Overwrite geometry (KML is authoritative).
        pin["id"] = culvert_id
        pin["geometry"] = {"type": "LineString", "coordinates": coords}
        pin["sta_ft"] = round(sta_ft, 2)
        pin["sta"] = _fmt_sta(sta_ft)
        # Overwrite XLSX-owned attrs.
        attrs = pin.setdefault("attrs", {})
        x = m["xlsx"]
        for k, v in {
            "material":            x["material"],
            "in_type":             x["in_type"],
            "in_side":             x["in_side"],
            "out_type":            x["out_type"],
            "out_side":            x["out_side"],
            "length_ft":           x["length_ft"],
            "size_in":             x["size_in"],
            "drainage_basin_ac":   x["drainage_basin_ac"],
            "drainage_direction":  x["drainage_direction"],
            "notes":               x["notes"],
            "stationing_label":    x["stationing_label"],
            "pla94_sta_label":     x["pla94_sta_label"],
            "survey_num":          m["kml"]["num"],
        }.items():
            if v == "" or v is None:
                attrs.pop(k, None)
            else:
                attrs[k] = v

    # ── Orphans: bundle pins in target sections not in new source ──
    # Split into "safe to delete" (empty legacy pins with no field
    # data) and "keep + flag" (still carry photos / conditions / notes
    # the user hasn't decided about). Only auto-drop the empty ones.
    orphans_dropped = []
    orphans_kept    = []
    for (sec_id, cid), pin in existing_by_sec_id.items():
        if (sec_id, cid) in seen_pin_keys:
            continue
        if _pin_score(pin) == 0:
            # No photos, no conditions, no notes — legacy stub.
            sec = sections_by_id.get(sec_id)
            if sec:
                sec["pins"] = [p for p in (sec.get("pins") or []) if p is not pin]
            orphans_dropped.append((sec_id, cid))
        else:
            orphans_kept.append((sec_id, cid))

    # ── Sort pins per section by sta_ft for stable diffs ────────
    for sec in sections_by_id.values():
        sec.get("pins", []).sort(
            key=lambda p: (p.get("sta_ft") if p.get("sta_ft") is not None else 0,
                           p.get("id") or "")
        )

    # ── Write ───────────────────────────────────────────────────
    BUNDLE_PATH.write_text(json.dumps(bundle, separators=(",", ":")) + "\n",
                           encoding="utf-8")

    # ── Summary ─────────────────────────────────────────────────
    print()
    if deduped_count:
        print(f"  Pre-dedupe: dropped  {deduped_count} duplicate bundle pin(s) "
              f"(kept the copy with more inspection data)")
    print(f"  Matched (KML+XLSX):  {len(matched)}")
    print(f"    Updated in place:  {updated}")
    print(f"    Created new pin:   {created}")
    if kml_only:
        print(f"  KML Num with no XLSX Order match:  {len(kml_only)}")
        for n in kml_only:
            print(f"     ! Num {n}")
    if xlsx_only:
        print(f"  XLSX Order with no KML Num match:  {len(xlsx_only)}")
        for n in xlsx_only:
            print(f"     ! Order {n}")
    if orphans_dropped:
        print(f"  Orphan pins auto-dropped (empty legacy stubs):  {len(orphans_dropped)}")
        for sec_id, cid in orphans_dropped:
            print(f"     - {sec_id} · {cid}")
    if orphans_kept:
        print(f"  Orphan pins KEPT (carry field data — user review):  {len(orphans_kept)}")
        for sec_id, cid in orphans_kept:
            print(f"     ~ {sec_id} · {cid}")
    print(f"\nWrote {BUNDLE_PATH}.")
    print("Photos: preserved from prior bundle state — this script does "
          "not touch attrs.photos / report_hidden_photos.")
    print("Elevations / bearing / drop / slope: preserved — re-run "
          "build_flow_geometry.py + finalize_geometry.py to refresh those "
          "against the new KML geometry.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
