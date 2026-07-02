"""Extract Excel "image-in-cell" (RichValue) photos from the JSG 070126
XLSX and emit a JSONL manifest that RoadWalk's existing "📋 Import
matched photos…" toolbar button consumes.

The XLSX stores photos via Excel's newer RichValue mechanism:
  Cell   vm="N"  →  xl/metadata.xml <bk> at N  →  xlrd:rvb i="I"
  rdrichvalue.xml <rv> at I → <v>K</v>
  richValueRel.xml <rel> at K → r:id="rIdM"
  _rels/richValueRel.xml.rels rIdM → xl/media/imageX.<ext>

We resolve that whole chain per cell in cols Q..T (Inlet Photo 1, 2 ·
Outlet Photo 1, 2), map row → (Section ID, Culvert ID, Order) via the
Sheet1 body cols, then downscale + watermark + re-encode each JPEG at
1200 px / q=82 (matching build_culvert_photos_manifest.py). Output is
JSONL — one header line then one record per photo — sized to survive
the streaming importer at scale (100+ MB safe).

Output path is next to the XLSX so the user can pick it via the
importer's file input without hunting.

Run from anywhere:
    python data/build_photos_manifest_from_xlsx_cells.py
"""

from __future__ import annotations

import base64
import io
import json
import re
import sys
import zipfile
from datetime import datetime
from pathlib import Path

try:
    from PIL import Image, ImageDraw, ImageFont, ImageOps
except ImportError:
    print("ERROR: install Pillow — `pip install pillow`", file=sys.stderr)
    sys.exit(1)

XLSX_PATH = (
    Path.home()
    / "OneDrive - AECOM" / "Documents" / "!DATA" / "EFL"
    / "2026 GRSM" / "GRSM_Culvert"
    / "GRSM_Culvert_Stationing_SF+JSG 070126 (original phtoos removed).xlsx"
)
OUT_PATH = XLSX_PATH.with_name("culvert_photos_from_xlsx_manifest.jsonl")

MAX_SIZE = 1200
JPEG_QUALITY = 82

SECTIONS_OF_INTEREST = {"A", "G", "H", "I"}

# Sheet1 body column layout — same as the KML+XLSX importer.
COL_SECTION_ID    = 0
COL_CULVERT_ID    = 1
COL_ORDER         = 2
# Photo columns are 0-indexed 16..19 (spreadsheet Q, R, S, T).
PHOTO_COL_ENDS = {
    16: ("inlet",  1),   # Q — Inlet Photo 1
    17: ("inlet",  2),   # R — Inlet Photo 2
    18: ("outlet", 1),   # S — Outlet Photo 1
    19: ("outlet", 2),   # T — Outlet Photo 2
}

# Regexes over the raw XML — dependency-free so a fresh clone runs
# without extra installs beyond Pillow.
_CELL_RE  = re.compile(r'<c r="([A-Z]+\d+)"[^>]*\bvm="(\d+)"[^>]*/?>')
_ROW_RE   = re.compile(r'<row r="(\d+)"[^>]*>([\s\S]*?)</row>')
_BK_RE    = re.compile(r'<bk>[\s\S]*?<xlrd:rvb i="(\d+)"/>[\s\S]*?</bk>')
_RV_RE    = re.compile(r'<rv[^>]*>[\s\S]*?<v>(\d+)</v>')
_REL_RE   = re.compile(r'<rel r:id="(rId\d+)"/>')
_RELSHIP_RE = re.compile(r'<Relationship\s+Id="([^"]+)"[^>]*Target="([^"]+)"')


# ── Column letter helpers ─────────────────────────────────────────
def _col_letter_to_idx(letters: str) -> int:
    """A → 0, B → 1, …, AA → 26, …"""
    n = 0
    for ch in letters:
        n = n * 26 + (ord(ch.upper()) - ord("A") + 1)
    return n - 1


def _cell_ref_to_rc(ref: str) -> tuple[int, int]:
    """'Q3' → (row_1based, col_0based)."""
    m = re.match(r"^([A-Z]+)(\d+)$", ref)
    if not m:
        raise ValueError(f"bad cell ref {ref!r}")
    return int(m.group(2)), _col_letter_to_idx(m.group(1))


# ── XLSX → cell/image lookup ─────────────────────────────────────
def _load_lookup_chain(z: zipfile.ZipFile) -> tuple[dict[int, str], dict[str, bytes]]:
    """Return (vm → media_zip_path, media_zip_path → bytes)."""
    # 1. metadata.xml — 144 <bk> entries. Each carries xlrd:rvb i.
    md = z.read("xl/metadata.xml").decode("utf-8", errors="replace")
    vm_to_rvb = {i: int(m.group(1))
                 for i, m in enumerate(_BK_RE.finditer(md))}
    # 2. rdrichvalue.xml — <rv> per rich value, first <v> is the rel key.
    rv_xml = z.read("xl/richData/rdrichvalue.xml").decode("utf-8", errors="replace")
    rvb_to_key = {i: int(m.group(1))
                  for i, m in enumerate(_RV_RE.finditer(rv_xml))}
    # 3. richValueRel.xml — <rel r:id="rIdN"/> per key, order-preserving.
    rel_xml = z.read("xl/richData/richValueRel.xml").decode("utf-8", errors="replace")
    key_to_rid = {i: m.group(1)
                  for i, m in enumerate(_REL_RE.finditer(rel_xml))}
    # 4. _rels/richValueRel.xml.rels — rId → media path.
    rels_xml = z.read("xl/richData/_rels/richValueRel.xml.rels").decode("utf-8", errors="replace")
    rid_to_target = dict(_RELSHIP_RE.findall(rels_xml))
    # Compose vm → media_zip_path.
    vm_to_path: dict[int, str] = {}
    for vm, rvb in vm_to_rvb.items():
        key = rvb_to_key.get(rvb)
        if key is None: continue
        rid = key_to_rid.get(key)
        if rid is None: continue
        tgt = rid_to_target.get(rid)
        if not tgt: continue
        # Target is relative to xl/richData/. Normalise back to xl/…
        p = f"xl/richData/{tgt}".replace("/./", "/")
        while "/../" in p:
            head, _, tail = p.partition("/../")
            p = head.rsplit("/", 1)[0] + "/" + tail
        vm_to_path[vm] = p
    return vm_to_path, {}


# ── Row + photo cell walk ────────────────────────────────────────
def _walk_photo_cells(z: zipfile.ZipFile) -> list[tuple[int, int, int]]:
    """Return [(row_1based, col_0based, vm), ...] for every photo-column
    cell that carries a vm attribute."""
    sheet = z.read("xl/worksheets/sheet1.xml").decode("utf-8", errors="replace")
    out: list[tuple[int, int, int]] = []
    for m in _CELL_RE.finditer(sheet):
        row, col = _cell_ref_to_rc(m.group(1))
        if col not in PHOTO_COL_ENDS:
            continue
        out.append((row, col, int(m.group(2))))
    return out


# ── XLSX body rows (Section / Culvert ID / Order) ────────────────
def _read_body_rows(xlsx_path: Path) -> dict[int, dict]:
    """row_1based → {section_id, culvert_id, order}"""
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    rows: dict[int, dict] = {}
    for i, r in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1: continue
        if not r: continue
        sec_id = _str(r[COL_SECTION_ID] if len(r) > COL_SECTION_ID else "")
        cid    = _str(r[COL_CULVERT_ID] if len(r) > COL_CULVERT_ID else "")
        order  = _str(r[COL_ORDER]      if len(r) > COL_ORDER      else "")
        # Section ID falls back to Culvert ID prefix when the cell is blank.
        if not sec_id and "-" in cid:
            sec_id = cid.split("-", 1)[0]
        if sec_id not in SECTIONS_OF_INTEREST:
            continue
        rows[i] = {"section_id": sec_id, "culvert_id": cid, "order": order}
    return rows


def _str(v) -> str:
    if v is None: return ""
    if isinstance(v, float) and v.is_integer(): v = int(v)
    return str(v).strip()


# ── Watermark ─────────────────────────────────────────────────────
def _wm_font(size_px: int):
    for c in ("C:/Windows/Fonts/consolab.ttf", "consolab.ttf",
              "DejaVuSansMono-Bold.ttf"):
        try:
            return ImageFont.truetype(c, size_px)
        except Exception:
            continue
    return ImageFont.load_default()


def _burn_watermark(img: Image.Image, lines: list[str]) -> Image.Image:
    if not lines: return img
    if img.mode != "RGB":
        img = img.convert("RGB")
    w, h = img.size
    fs = max(12, round(w / 48))
    font = _wm_font(fs)
    lh = round(fs * 1.5)
    pad_x = round(fs * 0.6)
    pad_y = round(fs * 0.45)
    band_h = len(lines) * lh + pad_y * 2
    overlay = Image.new("RGBA", img.size, (0, 0, 0, 0))
    draw = ImageDraw.Draw(overlay)
    draw.rectangle([(0, h - band_h), (w, h)], fill=(0, 0, 0, int(0.58 * 255)))
    for i, ln in enumerate(lines):
        draw.text((pad_x, h - band_h + pad_y + i * lh), ln,
                  font=font, fill=(255, 255, 255, 255))
    return Image.alpha_composite(img.convert("RGBA"), overlay).convert("RGB")


def _to_jpeg_dataurl(raw_bytes: bytes, wm_lines: list[str]) -> str:
    img = Image.open(io.BytesIO(raw_bytes))
    img = ImageOps.exif_transpose(img)
    if img.mode != "RGB":
        img = img.convert("RGB")
    w, h = img.size
    if max(w, h) > MAX_SIZE:
        s = MAX_SIZE / max(w, h)
        img = img.resize((max(1, round(w * s)), max(1, round(h * s))),
                         Image.LANCZOS)
    img = _burn_watermark(img, wm_lines)
    buf = io.BytesIO()
    img.save(buf, "JPEG", quality=JPEG_QUALITY, optimize=True)
    return "data:image/jpeg;base64," + base64.b64encode(buf.getvalue()).decode("ascii")


# ── Main ─────────────────────────────────────────────────────────
def main() -> int:
    if not XLSX_PATH.exists():
        print(f"ERROR: XLSX not found at {XLSX_PATH}", file=sys.stderr)
        return 1
    print(f"Reading XLSX  : {XLSX_PATH}")
    body_rows = _read_body_rows(XLSX_PATH)
    print(f"  body rows in target sections: {len(body_rows)}")
    with zipfile.ZipFile(XLSX_PATH) as z:
        vm_to_path, _ = _load_lookup_chain(z)
        print(f"  vm→media entries: {len(vm_to_path)}")
        photo_cells = _walk_photo_cells(z)
        print(f"  photo-column cells with vm: {len(photo_cells)}")

        processed: list[dict] = []
        skipped_no_row = 0
        skipped_no_media = 0
        failed_decode: list[str] = []
        seen_paths: set[str] = set()

        for row, col, vm in photo_cells:
            body = body_rows.get(row)
            if not body:
                skipped_no_row += 1; continue
            media_path = vm_to_path.get(vm)
            if not media_path or media_path not in z.namelist():
                skipped_no_media += 1; continue
            end, slot = PHOTO_COL_ENDS[col]
            sec_id = body["section_id"]; cid = body["culvert_id"]
            end_label = f"{end.title()} {slot}"
            dedupe_key = f"{sec_id}::{cid}::{end}::{slot}"
            if dedupe_key in seen_paths:
                continue
            seen_paths.add(dedupe_key)
            wm_lines = [
                f"{cid}",
                f"{end_label}",
                f"XLSX SF+JSG 070126",
            ]
            try:
                data_url = _to_jpeg_dataurl(z.read(media_path), wm_lines)
            except Exception as e:
                failed_decode.append(f"{cid} {end_label}: {e}")
                continue
            processed.append({
                "id":            f"xlsx_{sec_id}_{cid}_{end}_{slot}",
                "culvert_id":    cid,
                "section_id":    sec_id,
                "end":           end,                # "inlet" or "outlet"
                "photo_slot":    slot,               # 1 or 2 within that end
                "photo_set":     "XLSX SF+JSG 070126",
                "original_file": media_path.rsplit("/", 1)[-1],
                "captured_at":   "",
                "lat":           None,
                "lon":           None,
                "dist_m":        None,
                "dataUrl":       data_url,
            })
            if len(processed) % 20 == 0:
                print(f"    processed {len(processed)}…")

    # ── Write manifest ─────────────────────────────────────────
    culvert_ids = sorted({p["culvert_id"] for p in processed})
    header = {
        "format":         "culvert_photos_manifest_jsonl_v1",
        "exported_at":    datetime.now().isoformat(timespec="seconds"),
        "source_xlsx":    str(XLSX_PATH),
        "total":          len(processed),
        "culvert_ids":    culvert_ids,
        "image_settings": {
            "max_size_px":  MAX_SIZE,
            "jpeg_quality": JPEG_QUALITY,
            "watermark":    "bottom-band black 58% / white mono / "
                            "culvert_id · end · source",
        },
    }
    print()
    print(f"Writing manifest → {OUT_PATH}")
    with OUT_PATH.open("w", encoding="utf-8") as fh:
        fh.write(json.dumps(header, separators=(",", ":")) + "\n")
        for rec in processed:
            fh.write(json.dumps(rec, separators=(",", ":")) + "\n")
    size_mb = OUT_PATH.stat().st_size / 1_048_576
    print(f"  {len(processed)} photo(s) across {len(culvert_ids)} culvert(s) "
          f"— {size_mb:.1f} MB")
    if skipped_no_row:
        print(f"  skipped (row not in target sections): {skipped_no_row}")
    if skipped_no_media:
        print(f"  skipped (vm → media path resolution failed): {skipped_no_media}")
    if failed_decode:
        print(f"  failed decode: {len(failed_decode)}")
        for m in failed_decode[:10]:
            print(f"     ! {m}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
